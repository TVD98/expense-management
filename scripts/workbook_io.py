#!/usr/bin/env python3
"""Shared Excel helpers for expense-tracker (1 month file, 1 sheet per day)."""

from __future__ import annotations

import re
from datetime import date, datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation

# Day sheet columns (date is the sheet itself)
HEADERS = ["STT", "Số tiền", "Danh mục", "Ghi chú"]
LEGACY_HEADERS = ["Ngày", "Số tiền", "Danh mục", "Ghi chú"]
CATEGORIES = [
    "Ăn uống",
    "Di chuyển",
    "Nhà ở",
    "Mua sắm",
    "Sức khỏe",
    "Giải trí",
    "Khác",
]
# Soft but distinct fills (easy to scan; Nhà ở ≠ Sức khỏe)
CATEGORY_COLORS = {
    "Ăn uống": "FBE2D5",   # peach
    "Di chuyển": "D6EAF8",  # sky blue
    "Nhà ở": "D5F5E3",      # mint teal
    "Mua sắm": "FADBD8",    # soft rose
    "Sức khỏe": "D4EFDF",   # leaf green
    "Giải trí": "E8DAEF",   # soft violet
    "Khác": "EAECEE",       # cool gray
}

# Typography (macOS-friendly; Excel falls back if missing)
FONT_NAME = "Helvetica Neue"
FONT_TITLE = "Avenir Next"
HEADER_FILL = PatternFill("solid", fgColor="2C3E50")
HEADER_FONT = Font(bold=True, color="FFFFFF", name=FONT_NAME, size=11)
TITLE_FONT = Font(bold=True, name=FONT_TITLE, size=15, color="2C3E50")
DATA_FONT = Font(name=FONT_NAME, size=11, bold=False)
CAT_FONT = Font(name=FONT_NAME, size=11, bold=True, color="2C3E50")
TOTAL_FONT = Font(bold=True, name=FONT_NAME, size=11, color="2C3E50")
MONEY_FORMAT = '#,##0'
THIN = Border(
    left=Side(style="thin", color="D5D8DC"),
    right=Side(style="thin", color="D5D8DC"),
    top=Side(style="thin", color="D5D8DC"),
    bottom=Side(style="thin", color="D5D8DC"),
)
ALT_FILL = PatternFill("solid", fgColor="F2F2F2")
TOTAL_FILL = PatternFill("solid", fgColor="FCF3CF")
DEFAULT_DIR = Path.home() / "Documents" / "ChiTieu"

def parse_amount(raw: str) -> int:
    """Parse VND: 45k→45000, 1.5M→1500000, 45.000→45000."""
    cleaned = raw.strip().replace(" ", "")
    suffix_match = re.fullmatch(
        r"([+-]?\d+(?:[.,]\d+)?)\s*([kKmM]?)",
        cleaned,
        flags=re.IGNORECASE,
    )
    if not suffix_match:
        raise ValueError(f"Invalid amount: {raw!r}")

    num_str, suffix = suffix_match.group(1), suffix_match.group(2)
    if suffix:
        value = float(num_str.replace(",", "."))
        mult = 1_000 if suffix.lower() == "k" else 1_000_000
        amount = int(round(value * mult))
    else:
        digits = re.sub(r"[^\d]", "", cleaned)
        if not digits:
            raise ValueError(f"Invalid amount: {raw!r}")
        amount = int(digits)

    if amount <= 0:
        raise ValueError("Amount must be positive")
    return amount


def month_path(data_dir: Path, d: date) -> Path:
    return data_dir / f"chi-tieu-{d.year:04d}-{d.month:02d}.xlsx"


def sheet_name_for(d: date) -> str:
    """Sheet tab = day of month, zero-padded (01..31)."""
    return f"{d.day:02d}"


def parse_cell_date(value) -> date | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    s = str(value).strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def parse_amount_cell(value) -> int:
    if value is None:
        return 0
    if isinstance(value, (int, float)):
        return int(value)
    digits = "".join(ch for ch in str(value) if ch.isdigit())
    return int(digits) if digits else 0


def _apply_category_dropdown(ws) -> None:
    """Add/refresh dropdown for Danh mục column (C3:C200)."""
    # Remove previous category validations on this sheet to avoid duplicates
    keep = []
    for dv in list(ws.data_validations.dataValidation):
        ranges = str(dv.sqref) if dv.sqref else ""
        if "C3" in ranges or ranges.startswith("C") or ":C" in ranges:
            # drop old category list validations covering col C
            if dv.type == "list":
                continue
        keep.append(dv)
    ws.data_validations.dataValidation = keep

    formula = '"' + ",".join(CATEGORIES) + '"'
    dv = DataValidation(
        type="list",
        formula1=formula,
        allow_blank=True,
        showDropDown=False,  # False = show arrow (openpyxl quirk)
        showErrorMessage=True,
        errorTitle="Danh mục",
        error="Chọn một danh mục trong danh sách.",
        promptTitle="Danh mục",
        prompt="Chọn danh mục",
    )
    dv.add("C3:C200")
    ws.add_data_validation(dv)


def _style_day_sheet(ws, d: date) -> None:
    """Apply title, header, column widths, freeze panes."""
    title = f"Chi tiêu ngày {d.day:02d}/{d.month:02d}/{d.year}"
    if ws["A1"].value != title:
        # Only set title row if sheet looks new / wrong title
        if ws["A1"].value is None or str(ws["A1"].value).startswith("Chi tiêu"):
            ws["A1"] = title
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:D1")
    ws["A1"].alignment = Alignment(vertical="center", horizontal="left")
    ws.row_dimensions[1].height = 28

    # Ensure header row
    for col, h in enumerate(HEADERS, start=1):
        cell = ws.cell(2, col, h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN
    ws.row_dimensions[2].height = 20

    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 36
    ws.freeze_panes = "A3"
    ws.print_title_rows = "1:2"
    _apply_category_dropdown(ws)


def _data_end_row(ws) -> int:
    """Last data row index (before Tổng), or 2 if empty."""
    last = 2
    for r in range(3, (ws.max_row or 2) + 1):
        val = ws.cell(r, 1).value
        if val is None:
            continue
        if str(val).strip().lower() in ("tổng", "tong", "total"):
            break
        last = r
    return last


def _refresh_total_row(ws) -> None:
    end = _data_end_row(ws)
    # Clear old total rows below data
    for r in range(end + 1, (ws.max_row or end) + 2):
        for c in range(1, 5):
            ws.cell(r, c).value = None
            ws.cell(r, c).fill = PatternFill()
            ws.cell(r, c).font = Font()
            ws.cell(r, c).border = Border()

    total_row = end + 1
    if end < 3:
        # no data yet — still show total 0
        total_row = 3
        ws.cell(3, 1).value = None

    label = ws.cell(total_row, 1, "Tổng")
    label.font = TOTAL_FONT
    label.fill = TOTAL_FILL
    label.border = THIN

    if end >= 3:
        formula_cell = ws.cell(total_row, 2, f"=SUM(B3:B{end})")
    else:
        formula_cell = ws.cell(total_row, 2, 0)
    formula_cell.number_format = MONEY_FORMAT
    formula_cell.font = TOTAL_FONT
    formula_cell.fill = TOTAL_FILL
    formula_cell.border = THIN

    for c in (3, 4):
        cell = ws.cell(total_row, c, None)
        cell.fill = TOTAL_FILL
        cell.border = THIN


def _category_fill(category: str) -> PatternFill:
    hex_color = CATEGORY_COLORS.get((category or "").strip(), CATEGORY_COLORS["Khác"])
    return PatternFill("solid", fgColor=hex_color)


def _style_data_row(ws, row: int, stt: int, category: str | None = None) -> None:
    """Style a data row; color whole row by category."""
    if category is None:
        category = str(ws.cell(row, 3).value or "").strip()
    fill = _category_fill(category)
    for c in range(1, 5):
        cell = ws.cell(row, c)
        cell.font = DATA_FONT
        cell.border = THIN
        cell.fill = fill
    # Emphasize category cell
    cat_cell = ws.cell(row, 3)
    cat_cell.font = CAT_FONT

    ws.cell(row, 1).alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row, 2).number_format = MONEY_FORMAT
    ws.cell(row, 2).alignment = Alignment(horizontal="right", vertical="center")
    ws.cell(row, 3).alignment = Alignment(horizontal="left", vertical="center")
    ws.cell(row, 4).alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[row].height = 18


def ensure_day_sheet(wb, d: date):
    name = sheet_name_for(d)
    if name in wb.sheetnames:
        ws = wb[name]
    else:
        if len(wb.sheetnames) == 1 and wb.active.title == "Sheet" and wb.active.max_row == 1 and wb.active["A1"].value is None:
            ws = wb.active
            ws.title = name
        else:
            ws = wb.create_sheet(title=name)
        # Initialize layout
        ws["A1"] = f"Chi tiêu ngày {d.day:02d}/{d.month:02d}/{d.year}"
        for col, h in enumerate(HEADERS, start=1):
            ws.cell(2, col, h)

    _style_day_sheet(ws, d)
    return ws


def open_or_create_month_workbook(path: Path):
    path.parent.mkdir(parents=True, exist_ok=True)
    if path.exists():
        wb = load_workbook(path)
        migrate_legacy_if_needed(wb, path)
        return wb
    wb = Workbook()
    # leave default sheet; first add will rename/create day sheet
    return wb


def migrate_legacy_if_needed(wb, path: Path) -> bool:
    """Split old single-sheet 'Ngày|Số tiền|...' layout into day sheets."""
    legacy_sheets = []
    for name in list(wb.sheetnames):
        ws = wb[name]
        first = [ws.cell(1, c).value for c in range(1, 5)]
        if first == LEGACY_HEADERS or (
            first[0] == "Ngày" and first[1] == "Số tiền"
        ):
            legacy_sheets.append(ws)

    if not legacy_sheets:
        return False

    from collections import defaultdict

    by_day: dict[date, list[tuple]] = defaultdict(list)
    for ws in legacy_sheets:
        for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
            if i == 1:
                continue
            if not row or all(c is None for c in row):
                continue
            d = parse_cell_date(row[0])
            if d is None:
                continue
            amount = parse_amount_cell(row[1] if len(row) > 1 else 0)
            category = str(row[2] or "").strip() or "Khác"
            note = str(row[3] or "").strip() if len(row) > 3 else ""
            by_day[d].append((amount, category, note))

    # Remove legacy sheets
    for ws in legacy_sheets:
        wb.remove(ws)

    # If workbook empty, keep at least one sheet temporarily
    if not wb.sheetnames:
        wb.create_sheet("tmp")

    for d in sorted(by_day.keys()):
        ws = ensure_day_sheet(wb, d)
        # clear any leftover
        end = _data_end_row(ws)
        for r in range(3, max(end, 3) + 5):
            for c in range(1, 5):
                ws.cell(r, c).value = None
        for idx, (amount, category, note) in enumerate(by_day[d], start=1):
            row = 2 + idx
            ws.cell(row, 1, idx)
            ws.cell(row, 2, amount)
            ws.cell(row, 3, category)
            ws.cell(row, 4, note)
            _style_data_row(ws, row, idx)
        _refresh_total_row(ws)
        _style_day_sheet(ws, d)

    if "tmp" in wb.sheetnames and len(wb.sheetnames) > 1:
        wb.remove(wb["tmp"])

    # Sort sheets by day number
    order = sorted(
        [s for s in wb.sheetnames if re.fullmatch(r"\d{2}", s)],
        key=lambda s: int(s),
    )
    for i, name in enumerate(order):
        wb.move_sheet(name, offset=i - wb.sheetnames.index(name))

    wb.save(path)
    return True


def append_expense(path: Path, d: date, amount: int, category: str, note: str) -> dict:
    wb = open_or_create_month_workbook(path)
    ws = ensure_day_sheet(wb, d)

    end = _data_end_row(ws)
    if end < 3:
        next_stt = 1
        row = 3
    else:
        next_stt = end - 2 + 1
        row = end + 1

    ws.cell(row, 1, next_stt)
    ws.cell(row, 2, amount)
    ws.cell(row, 3, category)
    ws.cell(row, 4, note)
    _style_data_row(ws, row, next_stt, category)
    _refresh_total_row(ws)
    _style_day_sheet(ws, d)

    # Remove leftover default "Sheet" if empty and not our day
    for name in list(wb.sheetnames):
        if name == "Sheet" and name != sheet_name_for(d):
            other = wb[name]
            if other.max_row == 1 and other["A1"].value is None:
                wb.remove(other)

    # Keep day sheets sorted
    order = sorted(
        [s for s in wb.sheetnames if re.fullmatch(r"\d{2}", s)],
        key=lambda s: int(s),
    )
    for i, name in enumerate(order):
        current = wb.sheetnames.index(name)
        if current != i:
            wb.move_sheet(name, offset=i - current)

    wb.save(path)
    return {
        "ok": True,
        "file": str(path),
        "sheet": sheet_name_for(d),
        "date": d.isoformat(),
        "amount": amount,
        "category": category,
        "note": note,
        "stt": next_stt,
    }


def load_day_rows(path: Path, d: date | None = None) -> list[dict]:
    """Load expense rows. If d is set, only that day sheet; else all day sheets."""
    if not path.exists():
        return []
    wb = load_workbook(path, data_only=True)
    # migrate needs write — if legacy, reload after migrate with write workbook
    wb_check = load_workbook(path)
    if migrate_legacy_if_needed(wb_check, path):
        wb = load_workbook(path, data_only=True)

    rows: list[dict] = []
    year_month = None
    # Infer year-month from filename chi-tieu-YYYY-MM.xlsx
    m = re.search(r"chi-tieu-(\d{4})-(\d{2})", path.name)
    if m:
        year_month = (int(m.group(1)), int(m.group(2)))

    sheets = wb.sheetnames
    if d is not None:
        name = sheet_name_for(d)
        sheets = [name] if name in wb.sheetnames else []

    for name in sheets:
        if not re.fullmatch(r"\d{2}", name):
            # try legacy single sheet fallback
            ws = wb[name]
            first = [ws.cell(1, c).value for c in range(1, 5)]
            if first == LEGACY_HEADERS:
                for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
                    if i == 1:
                        continue
                    if not row or all(c is None for c in row):
                        continue
                    dd = parse_cell_date(row[0])
                    if dd is None:
                        continue
                    if d is not None and dd != d:
                        continue
                    rows.append(
                        {
                            "date": dd.isoformat(),
                            "amount": parse_amount_cell(row[1]),
                            "category": str(row[2] or "").strip() or "Khác",
                            "note": str(row[3] or "").strip() if len(row) > 3 else "",
                        }
                    )
            continue

        day_num = int(name)
        if year_month is None:
            continue
        y, mo = year_month
        try:
            sheet_date = date(y, mo, day_num)
        except ValueError:
            continue
        if d is not None and sheet_date != d:
            continue

        ws = wb[name]
        for r in range(3, (ws.max_row or 2) + 1):
            stt = ws.cell(r, 1).value
            if stt is None:
                continue
            if str(stt).strip().lower() in ("tổng", "tong", "total"):
                break
            amount = parse_amount_cell(ws.cell(r, 2).value)
            if amount <= 0:
                continue
            rows.append(
                {
                    "date": sheet_date.isoformat(),
                    "amount": amount,
                    "category": str(ws.cell(r, 3).value or "").strip() or "Khác",
                    "note": str(ws.cell(r, 4).value or "").strip(),
                    "stt": stt,
                }
            )
    return rows
